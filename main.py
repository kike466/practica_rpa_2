import smtplib
import time
from email import encoders
from email.mime.base import MIMEBase
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText

from openpyxl import load_workbook
from selenium.webdriver import ActionChains, Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

from usuarios import usuario
from selenium import webdriver

import glob
import os



#Se abre el libro del excel y se extraen el DNI y la matriula
wb = load_workbook(filename='suma.xlsx')
sheet_ranges = wb['Suma']

usrs = []
for c1, c2 in sheet_ranges['A2':'B13']:
    DNI = c1.value
    Matricula = c2.value

    #Se crea un usuario
    usr = usuario(DNI, Matricula)
    #Se añade a una lista
    usrs.append(usr)
#cerramos el libro
wb.close()

#iniciamos el driver de chromedriver
driver = webdriver.Chrome(r'chromedriver_win32\chromedriver.exe')
#maximizamos la ventana
driver.maximize_window()
#accedemos a una url
driver.get("https://www.suma.es/")
time.sleep(1)

#action chains nos permite interactuar con los elementos
achain = ActionChains(driver)
#hacemos el hover del contribuyentes
search=WebDriverWait(driver,10).until(
    EC.presence_of_element_located(
        (By.XPATH, '//*[@id="navbarNavDropdown"]/ul/li[1]')
   )
)

achain.move_to_element(search).click().perform()

time.sleep(1)
# click en obtener recibo
driver.find_element(By.LINK_TEXT, 'Obtener un recibo').click()
time.sleep(1)
# click en Con el numero fijo
driver.find_element(By.LINK_TEXT, 'Con el número fijo (periodo voluntario)').click()
#se ban buscando resultados para cada usuario
resultado=[]


def obtener_ultimo_fichero():
    # obtener el ultimo fichero descargado
    list_of_files = glob.glob(r'C:\Users\Usuario\Downloads\*.pdf')  # * means all if need specific format then *.csv
    latest_file = max(list_of_files, key=os.path.getctime)
    print("Ultima descarga")
    print(latest_file)
    return latest_file

message = MIMEMultipart()
#email que envia
sender =''
#email que recive
receiver =''

message['From'] = sender
message['To'] = receiver
message['Subject'] = 'This email has an attacment, a pdf file'
def crear_email(latest_file):
    # Setup the MIME


    binary_pdf = open(latest_file, 'rb')

    payload = MIMEBase('application', 'octate-stream', Name=latest_file)
    payload.set_payload((binary_pdf).read())

    # enconding the binary into base64
    encoders.encode_base64(payload)

    # add header with pdf name
    payload.add_header('Content-Decomposition', 'attachment', filename=latest_file)
    message.attach(payload)

def enviar_email():
    session = smtplib.SMTP('smtp.gmail.com', 587)

    # enable security
    session.starttls()

    #password del email que envia
    password = ''
    session.login(sender, password)

    text = message.as_string()
    session.sendmail(sender, receiver, text)
    session.quit()
    print('Mail Sent')



for usuario in usrs:
    #hacemos el clear de los input
    driver.find_element(By.ID, 'pantalla:nif').clear()
    driver.find_element(By.ID, 'pantalla:cprv_numerofijo').clear()
    #añadimos datos a los inputs
    elemento_DNI = driver.find_element(By.ID, 'pantalla:nif')
    elemento_Matricula = driver.find_element(By.ID, 'pantalla:cprv_numerofijo')

    elemento_DNI.send_keys(usuario.get_DNI())
    elemento_Matricula.send_keys(usuario.get_Matricula())

    time.sleep(1)
    #click en el enviar
    driver.find_element(By.ID, 'pantalla:obtenerRecibos').click()
    time.sleep(1)

    msg=''
#error de no se encuentra
    try:
        if driver.find_element(By.XPATH, '//*[@id="pantalla"]/div/div[4]/div[2]/div/span').is_displayed():
            msg='Error '+ driver.find_element(By.XPATH, '//*[@id="pantalla"]/div/div[4]/div[2]/div/span').text
            resultado.append(msg)
            print(msg)
    except:
        print("An exception occurred 1")
#error de que los datos introducidos no cumplen los requisitos
    try:

        if driver.find_element(By.XPATH, '//*[@id="pantalla"]/div/div[4]/div/div[1]/div[2]/span').is_displayed():
            msg = 'Error ' + driver.find_element(By.XPATH, '//*[@id="pantalla"]/div/div[4]/div/div[1]/div[2]/span').text
            resultado.append(msg)
            print(msg)

    except:
        print("An exception occurred 2")

#en caso de que todo este bien

    if msg=='':
            msg = 'OK'
            print(msg)
            resultado.append(msg)
            time.sleep(1)
            #descargar pdf
            driver.find_element(By.XPATH, '//*[@id="pantalla:listaVallistaValores"]/tbody/tr/td[9]/div/a').click()

            driver.find_element(By.XPATH, '//*[@id="panelbotones"]/div[3]/img').click()
            time.sleep(4)
            search = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located(
                    (By.XPATH, '//*[@id="pantalla:tabladocumentosGroup"]/table[1]/tbody/tr/td[3]/a/i')
                )
            )
            search.click()
            time.sleep(2)

            #obtenemos el ultimo fichero descargado
            latest_file = obtener_ultimo_fichero()
            #creamos el email
            crear_email(latest_file)

#enviamos el email
enviar_email()


#Se abre el libro para poner los mensajes de resultado
wb = load_workbook(filename='suma.xlsx')
sheet_ranges = wb['Suma']
count=2
print("--------------------------------")
for items in resultado:
    #se insertan los valores de los resultados
    sheet_ranges['C'+str(count)]=items
    count+=1

#se guardan los cambios
wb.save('suma.xlsx')
#se cierra el libro
wb.close()
#se cierra el chromedriver
driver.close()


